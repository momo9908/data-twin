# -*- coding: utf-8 -*-
"""
dic_analysis.py
===============
DIC 全场 Mises 等效应变离线分析模块 (纯计算, 无 PyQt 依赖)。

职责:
    从 GOM/Aramis 风格导出的 DIC 多时刻 Excel 中, 找出轮盘在所有时刻、全场
    范围内的 **最大 Mises 等效应变** 及其所在测点坐标。

    (历史上曾支持多种指标; 现按需求仅保留 Mises 等效应变一种分析。指标仍以
     注册表 _METRICS 形式组织, 便于将来按需扩展。)

    数据来源约定 (DICdata/20260520-DIC.xlsx):
        - 工作簿含 501 个工作表: STAGE000, STAGE001, ... , STAGE500
        - STAGE000 为零位参考态 (位移/应变全为 0), 不计入"时刻"统计
        - STAGE001 ~ STAGE500 为 500 个不同时刻 (轮盘 500RPM 下) 的全场数据
        - 每个工作表第 1 行为列标题, 其后每行是一个测点
        - 关键列 (按列标题文本稳健匹配, 不写死索引):
              行ROW | 列COL | X坐标[mm] | Y坐标[mm] | ... | Mises应变[%] | ...
        - 网格边缘测点 (无法由邻域差分出应变) 的应变列为空 (None), 需跳过

依赖:
    - openpyxl (read_only 流式读取, 避免 169MB 文件一次性吃满内存)
    - 标准库 dataclasses / glob / os / sys / time

    另含两类"截面平均"分析 (可选, 把点转圆柱坐标后求平均):
        - 平均周向Mises应变: 子午截面 (角度 θ 固定, 沿半径) 上各点 Mises 的平均
        - 平均径向Mises应变: 圆柱截面 (半径 r=R 固定, 绕圆周) 上各点 Mises 的平均
      二者均逐 sheet 计算, 跨全部时刻取最大值。

被调用方:
    - 命令行: ``python dic_analysis.py [xlsx路径] [半径R] [角度θ]`` 直接打印报告
    - GUI:    dic_tab.DicAnalysisTab._run_module → analyze_max_field
"""

import os
import sys
import glob
import time
import math
from dataclasses import dataclass, field
from typing import Callable, Optional

try:
    import openpyxl
except ImportError:  # pragma: no cover - 仅在缺依赖的环境触发
    openpyxl = None


# DIC 数据默认所在子目录 (相对程序所在目录)
DEFAULT_DIC_DIRNAME = 'DICdata'

# 截面平均的默认容差半带宽 (用于把散点归入"某截面"; 可在 analyze_max_field 覆盖)
DEFAULT_THETA_HALFBAND_DEG = 1.5   # 子午截面 θ 半带宽 (度, ≈一列点的角间距)
DEFAULT_RADIUS_HALFBAND = 1500.0   # 圆柱截面 r 半带宽 (mm, ≈半个网格步距)


# =============================================================================
# 分析指标 (模块) 注册表
# =============================================================================
# 每个指标: key -> (中文名, 单位, 列标题匹配函数)
# 列匹配函数对单元格标题文本返回 True 即认定为该指标列。
# 注: 当前仅保留 Mises 等效应变一种 (按"全场取最大值"语义统计)。
_METRICS = {
    'mises': ('Mises 等效应变', '%', lambda s: 'mises' in s.lower()),
}


def available_metrics():
    """返回所有可选分析指标 (供 UI 构建模块按钮)。

    Args:
        无

    Returns:
        list[tuple(key, label, unit)], 顺序即 UI 推荐展示顺序。
        当前仅含 Mises 一项。

    Side Effects:
        无 (纯函数)
    """
    return [(k, v[0], v[1]) for k, v in _METRICS.items()]


def metric_label(metric_key: str) -> str:
    """返回指标的中文名 (无效 key 回退为 key 本身)。

    Args:
        metric_key: 指标键 (如 'mises')

    Returns:
        中文名字符串

    Side Effects:
        无 (纯函数)
    """
    info = _METRICS.get(metric_key)
    return info[0] if info else metric_key


# =============================================================================
# 结果数据结构
# =============================================================================
@dataclass
class PointResult:
    """单个测点的指标结果 (用于"全场最大点")。

    字段:
        stage_name:  工作表名 (如 'STAGE427')
        stage_index: 时刻序号 (STAGE000=0 参考态, STAGE001=1, ...; 即表名数字)
        value:       该点的指标值 (单位见 DicFieldResult.unit)
        grid_row:    测点网格行号 (行ROW), 用于唯一标识物理测点
        grid_col:    测点网格列号 (列COL)
        x_mm:        该时刻 (变形后) 的 X 坐标, 单位 mm
        y_mm:        该时刻 (变形后) 的 Y 坐标, 单位 mm
        z_mm:        该时刻 (变形后) 的 Z 坐标, 单位 mm (圆柱坐标 z 分量); 缺失为 None
        ref_x_mm:    零位参考态 (STAGE000) 同一测点的 X 坐标 (mm); 缺失为 None
        ref_y_mm:    零位参考态同一测点的 Y 坐标 (mm); 缺失为 None
        ref_z_mm:    零位参考态同一测点的 Z 坐标 (mm); 缺失为 None
        excel_row:   该点在所属工作表中的行号 (含表头, 1-based), 便于回查原表
    """
    stage_name: str
    stage_index: int
    value: float
    grid_row: object
    grid_col: object
    x_mm: float
    y_mm: float
    z_mm: Optional[float] = None
    ref_x_mm: Optional[float] = None
    ref_y_mm: Optional[float] = None
    ref_z_mm: Optional[float] = None
    excel_row: int = 0


@dataclass
class SectionAvgResult:
    """某一截面上 Mises 平均值的"跨时刻最大"结果。

    字段:
        label:       显示名 ('平均周向Mises应变' 或 '平均径向Mises应变')
        value:       该最大值 (= 取得最大的那个 sheet 在该截面上的平均, 单位同
                     DicFieldResult.unit)
        stage_name:  取得最大值的工作表名
        stage_index: 时刻序号
        n_points:    该 sheet 在此截面内参与平均的测点数
        param_desc:  截面参数描述 (如 'θ=30.0°±1.5°' / 'r=90000±1500 mm')
    """
    label: str
    value: float
    stage_name: str
    stage_index: int
    n_points: int
    param_desc: str


@dataclass
class DicFieldResult:
    """整份 DIC 文件、某一指标的分析汇总 (仅保留全场最大点)。

    字段:
        file_path:        分析的 xlsx 绝对路径
        metric_key:       分析指标键 (如 'mises')
        metric_label:     指标中文名 (如 'Mises 等效应变')
        unit:             指标单位 (如 '%')
        n_stages_scanned: 实际参与"最大值"搜索的时刻数 (= 参考态以外的工作表数)
        reference_stage:  作为零位参考态的工作表名 (默认首表 STAGE000)
        global_max:       全场全时刻的最大值测点 (PointResult); 无有效数据时 None
        circ_section:     "平均周向Mises应变"(子午截面 θ 固定)的跨时刻最大; 未请求/
                          无数据时 None
        radial_section:   "平均径向Mises应变"(圆柱截面 r=R)的跨时刻最大; 未请求/
                          无数据时 None
        circ_theta_deg:   请求的子午截面角度 θ0 (度); None 表示未请求该项
        radial_radius:    请求的圆柱截面半径 R (mm); None 表示未请求该项
        theta_halfband_deg / radius_halfband: 两截面所用的容差半带宽 (回显)
        column_index:     实测的指标列索引 (0-based), 供排错
        elapsed_sec:      分析耗时 (秒)
        skipped_stages:   读取中途抛异常 (如坏单元格坐标) 被跳过的时刻表名列表;
                          被跳过表的截面平均不计入, 但其抛错前已读入的合法点仍
                          参与全场最大。空列表表示无跳过。
    """
    file_path: str
    metric_key: str
    metric_label: str
    unit: str
    n_stages_scanned: int
    reference_stage: str
    global_max: Optional[PointResult]
    circ_section: Optional[SectionAvgResult] = None
    radial_section: Optional[SectionAvgResult] = None
    circ_theta_deg: Optional[float] = None
    radial_radius: Optional[float] = None
    theta_halfband_deg: float = DEFAULT_THETA_HALFBAND_DEG
    radius_halfband: float = DEFAULT_RADIUS_HALFBAND
    column_index: int = -1
    elapsed_sec: float = 0.0
    skipped_stages: list = field(default_factory=list)


# =============================================================================
# 路径与列定位辅助
# =============================================================================
def _exe_dir() -> str:
    """返回程序所在目录的绝对路径。

    Args:
        无

    Returns:
        sys.argv[0] 所在目录的绝对路径

    Side Effects:
        无 (纯路径计算)

    实现说明:
        与 sim_data_source._exe_dir / settings_io._exe_dir 等价, 保持全项目
        在源码运行与 PyInstaller 打包后路径行为一致。
    """
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def list_dic_files(search_dir: str = None) -> list:
    """列出 DICdata 目录中所有可用的 DIC Excel 文件 (供 UI 下拉选择)。

    Args:
        search_dir: 待搜索目录; 默认 <exe_dir>/DICdata

    Returns:
        .xlsx 绝对路径列表 (按文件名升序); 目录不存在或无文件时返回 []。

    Side Effects:
        无 (仅列目录)

    算法不变量:
        - 必须忽略 Excel 打开文件时生成的临时锁文件 '~$*.xlsx'。
    """
    if search_dir is None:
        search_dir = os.path.join(_exe_dir(), DEFAULT_DIC_DIRNAME)
    if not os.path.isdir(search_dir):
        return []
    out = []
    for p in glob.glob(os.path.join(search_dir, '*.xlsx')):
        if os.path.basename(p).startswith('~$'):
            continue
        out.append(p)
    out.sort()
    return out


def find_dic_file(search_dir: str = None) -> Optional[str]:
    """在 DICdata 目录中自动定位首个 DIC 数据 Excel 文件。

    Args:
        search_dir: 待搜索目录; 默认 <exe_dir>/DICdata

    Returns:
        首个 .xlsx 绝对路径; 无可用文件返回 None。

    Side Effects:
        无 (仅列目录)
    """
    files = list_dic_files(search_dir)
    return files[0] if files else None


def _locate_columns(header: tuple, metric_match) -> dict:
    """按列标题文本稳健定位所需列的 0-based 索引 (不写死列号)。

    Args:
        header:       工作表首行 (列标题) 元组, 元素可能含 None
        metric_match: 指标列匹配函数 (对标题文本返回 bool)

    Returns:
        dict, 含键 'row' / 'col' / 'x' / 'y' / 'z' / 'metric', 值为对应列索引。
        其中 'z' 为可选 (找不到时为 -1, 圆柱坐标 z 分量按 0 处理)。

    Raises:
        ValueError: 当指标列或 X/Y 坐标列缺失时抛出, 提示数据格式不符。

    Side Effects:
        无 (纯解析)

    匹配规则 (取第一个命中列):
        - metric: 由调用方传入的 metric_match 决定 (见 _METRICS)
        - row:    标题含 'ROW' (对应 '行ROW')
        - col:    标题含 'COL' (对应 '列COL')
        - x:      标题含 'X坐标'
        - y:      标题含 'Y坐标'
        - z:      标题含 'Z坐标' (取第一个; 可选)
          注意: 不能只匹配 '坐标', 否则会先命中重复出现的 'Z坐标' 列。
    """
    def _find(predicate) -> int:
        for idx, h in enumerate(header):
            if h is None:
                continue
            if predicate(str(h)):
                return idx
        return -1

    idx_metric = _find(metric_match)
    idx_row = _find(lambda s: 'ROW' in s.upper())
    idx_col = _find(lambda s: 'COL' in s.upper())
    idx_x = _find(lambda s: 'X坐标' in s)
    idx_y = _find(lambda s: 'Y坐标' in s)
    idx_z = _find(lambda s: 'Z坐标' in s)   # 可选, 圆柱坐标 z 分量

    missing = [name for name, i in
               (('指标列', idx_metric), ('X坐标', idx_x), ('Y坐标', idx_y))
               if i < 0]
    if missing:
        raise ValueError(f'DIC 数据缺少必需列: {", ".join(missing)}; 实际表头={header}')

    return {'row': idx_row, 'col': idx_col, 'x': idx_x, 'y': idx_y, 'z': idx_z,
            'metric': idx_metric}


# =============================================================================
# 主分析
# =============================================================================
def analyze_max_field(
    path: str = None,
    metric_key: str = 'mises',
    reference_index: int = 0,
    progress_cb: Optional[Callable[[int, int], bool]] = None,
    circ_theta_deg: Optional[float] = None,
    radial_radius: Optional[float] = None,
    theta_halfband_deg: float = DEFAULT_THETA_HALFBAND_DEG,
    radius_halfband: float = DEFAULT_RADIUS_HALFBAND,
) -> DicFieldResult:
    """遍历所有时刻, 找出全场指定指标的最大值及其测点坐标; 并(可选)在同一次扫描
    中计算"截面平均"的跨时刻最大值。

    截面平均 (把每个点按 (r=√(x²+y²), θ=atan2(y,x)°) 归入截面后求平均):
        - 平均周向Mises应变: 子午截面 (角度 θ≈circ_theta_deg) 上各点平均;
          每个 sheet 得一个平均, 跨全部时刻取最大 → 结果 circ_section。
        - 平均径向Mises应变: 圆柱截面 (半径 r≈radial_radius) 上各点平均;
          每个 sheet 得一个平均, 跨全部时刻取最大 → 结果 radial_section。
      (命名沿用调用方约定: 周向↔子午截面、径向↔圆柱截面。)

    Args:
        path: DIC xlsx 绝对路径; None 时调用 find_dic_file() 自动定位。
        metric_key: 分析指标键, 见 _METRICS (默认且当前仅 'mises')。
        reference_index: 作为零位参考态的工作表序号 (默认 0, 即首表 STAGE000)。
            该表只用于提供各测点的"未变形参考坐标", 不参与最大值搜索。
        progress_cb: 可选进度回调 progress_cb(done_stages, total_stages)。
            - 每扫完一个时刻调用一次;
            - 若返回 False, 视为用户取消, 函数立即抛出 KeyboardInterrupt。
        circ_theta_deg: 子午截面角度 θ0 (度); None 则不计算"平均周向"。
        radial_radius:  圆柱截面半径 R (mm); None 则不计算"平均径向"。
        theta_halfband_deg: 子午截面 θ 容差半带宽 (度), 默认 DEFAULT_THETA_HALFBAND_DEG。
        radius_halfband:    圆柱截面 r 容差半带宽 (mm), 默认 DEFAULT_RADIUS_HALFBAND。

    Returns:
        DicFieldResult: 含全场最大点 global_max、以及 circ_section / radial_section
        (各为跨时刻最大的截面平均; 未请求或无数据点时为 None)。当文件无任何有效
        (非空) 指标数据时, global_max 为 None。

    Raises:
        RuntimeError: openpyxl 未安装 / path 未得到文件 / metric_key 无效。
        ValueError:   表头缺少必需列 (由 _locate_columns 抛出)。
        KeyboardInterrupt: progress_cb 返回 False (用户取消)。

    Side Effects:
        - 只读打开并流式遍历 xlsx (read_only=True, data_only=True), 不写盘、
          不修改任何全局状态; 结束时关闭工作簿。
        - 无日志打印 (打印交由 CLI / GUI 调用方, 保持本函数纯净)。

    算法不变量:
        - 指标列、坐标列、网格列一律按列标题文本定位, 不写死索引。
        - 参考态工作表 (reference_index) 不计入 n_stages_scanned, 也不参与
          最大值比较。
        - 网格边缘测点指标为 None 时必须跳过, 不可当作 0 参与比较。
        - "全场最大"= 所有被扫描时刻中、所有测点里指标值的最大者;
          并列时取先遇到的 (时刻序号小、行号小者)。
        - 某时刻表读取中途抛异常 (如坏单元格坐标 'W2W9') 时跳过该表余下部分:
          其截面平均因累加不完整而不予提交; 抛错前已读入的合法点仍参与全场最大;
          被跳过的表名记入返回结果的 skipped_stages, 不中断整份分析。
    """
    if openpyxl is None:
        raise RuntimeError('未安装 openpyxl, 无法读取 DIC Excel; 请先 pip install openpyxl')

    info = _METRICS.get(metric_key)
    if info is None:
        raise RuntimeError(f'未知分析指标: {metric_key!r}; 可选: {list(_METRICS)}')
    m_label, m_unit, m_match = info

    if path is None:
        path = find_dic_file()
    if not path or not os.path.exists(path):
        raise RuntimeError(f'未找到 DIC 数据文件 (path={path!r})')

    t0 = time.time()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            raise RuntimeError('DIC 工作簿没有任何工作表')

        ref_name = sheet_names[reference_index]
        instant_names = [n for i, n in enumerate(sheet_names) if i != reference_index]
        skipped = []   # 读取中途抛异常被跳过的时刻表名 (其截面平均不计入)

        # ---- 先用参考态首行定位列, 并构建 (行,列)->参考坐标 映射 ----
        ref_ws = wb[ref_name]
        cols = None
        ref_coord = {}
        try:
            for r_i, row in enumerate(ref_ws.iter_rows(values_only=True)):
                if r_i == 0:
                    cols = _locate_columns(row, m_match)   # 缺列会抛 ValueError, 须上抛
                    continue
                key = (row[cols['row']], row[cols['col']])
                zref = row[cols['z']] if cols['z'] >= 0 else None
                ref_coord[key] = (row[cols['x']], row[cols['y']], zref)
        except Exception:
            # 列尚未定位即抛 = 文件格式问题 (缺必需列), 原样上抛, 勿误吞为"空表"。
            # 已能定位列后才抛 = 个别参考坐标读不到, 保留已建映射继续 (下游对缺失
            # 坐标已按 None 兜底)。
            if cols is None:
                raise
            skipped.append(f'{ref_name}(参考态)')
        if cols is None:
            raise RuntimeError('参考态工作表为空, 无法定位列')

        ci_row, ci_col = cols['row'], cols['col']
        ci_x, ci_y, ci_z, ci_metric = cols['x'], cols['y'], cols['z'], cols['metric']
        # 只解析到所需最右列即可 (轻微提速)
        _cand = [ci_row, ci_col, ci_x, ci_y, ci_metric]
        if ci_z >= 0:
            _cand.append(ci_z)
        max_col_1based = max(_cand) + 1

        global_max: Optional[PointResult] = None
        circ_best: Optional[SectionAvgResult] = None      # 平均周向 (子午截面 θ)
        radial_best: Optional[SectionAvgResult] = None    # 平均径向 (圆柱截面 r=R)
        do_circ = circ_theta_deg is not None
        do_radial = radial_radius is not None
        need_sec = do_circ or do_radial
        total = len(instant_names)

        for si, name in enumerate(instant_names, start=1):
            ws = wb[name]
            c_sum = 0.0
            c_n = 0          # 子午截面 (θ≈θ0) 累加
            r_sum = 0.0
            r_n = 0          # 圆柱截面 (r≈R) 累加
            sheet_ok = True  # 本表是否完整读完 (中途抛异常则为 False)
            try:
                for r_i, row in enumerate(
                        ws.iter_rows(min_row=2, max_col=max_col_1based, values_only=True)):
                    v = row[ci_metric]
                    if v is None:
                        continue
                    x = row[ci_x]
                    y = row[ci_y]
                    # —— 截面平均: 把点按圆柱坐标 (r, θ) 归入对应截面 ——
                    if need_sec and x is not None and y is not None:
                        rr = math.hypot(float(x), float(y))
                        th = math.degrees(math.atan2(float(y), float(x)))
                        if do_circ and abs(th - circ_theta_deg) <= theta_halfband_deg:
                            c_sum += v
                            c_n += 1
                        if do_radial and abs(rr - radial_radius) <= radius_halfband:
                            r_sum += v
                            r_n += 1
                    # —— 全场最大点 ——
                    if global_max is None or v > global_max.value:
                        grow, gcol = row[ci_row], row[ci_col]
                        rx, ry, rz = ref_coord.get((grow, gcol), (None, None, None))
                        zval = row[ci_z] if ci_z >= 0 else None
                        global_max = PointResult(
                            stage_name=name,
                            stage_index=_stage_index(name),
                            value=float(v),
                            grid_row=grow,
                            grid_col=gcol,
                            x_mm=_as_float(x),
                            y_mm=_as_float(y),
                            z_mm=_as_float(zval),
                            ref_x_mm=_as_float(rx),
                            ref_y_mm=_as_float(ry),
                            ref_z_mm=_as_float(rz),
                            excel_row=r_i + 2,   # min_row=2 起算, 加回表头偏移
                        )
            except Exception:
                # 坏单元格坐标 (如非法引用 'W2W9') 等解析异常: openpyxl 的流式生成器
                # 一旦抛错即终止且无法续读, 故只能跳过本表余下部分。抛错前已读入的
                # 合法点已计入 global_max (予以保留); 本表截面累加不完整, 下方不提交。
                sheet_ok = False
                skipped.append(name)

            # —— 本 sheet 截面平均, 仅在整表成功读完时提交 (部分累加会偏, 不计) ——
            if sheet_ok and do_circ and c_n > 0:
                cavg = c_sum / c_n
                if circ_best is None or cavg > circ_best.value:
                    circ_best = SectionAvgResult(
                        label='平均周向Mises应变', value=cavg,
                        stage_name=name, stage_index=_stage_index(name), n_points=c_n,
                        param_desc=f'θ={circ_theta_deg:.1f}°±{theta_halfband_deg:.1f}°')
            if sheet_ok and do_radial and r_n > 0:
                ravg = r_sum / r_n
                if radial_best is None or ravg > radial_best.value:
                    radial_best = SectionAvgResult(
                        label='平均径向Mises应变', value=ravg,
                        stage_name=name, stage_index=_stage_index(name), n_points=r_n,
                        param_desc=f'r={radial_radius:.0f}±{radius_halfband:.0f} mm')

            if progress_cb is not None:
                if progress_cb(si, total) is False:
                    raise KeyboardInterrupt('用户取消 DIC 分析')

        return DicFieldResult(
            file_path=os.path.abspath(path),
            metric_key=metric_key,
            metric_label=m_label,
            unit=m_unit,
            n_stages_scanned=total,
            reference_stage=ref_name,
            global_max=global_max,
            circ_section=circ_best,
            radial_section=radial_best,
            circ_theta_deg=circ_theta_deg,
            radial_radius=radial_radius,
            theta_halfband_deg=theta_halfband_deg,
            radius_halfband=radius_halfband,
            column_index=ci_metric,
            elapsed_sec=time.time() - t0,
            skipped_stages=skipped,
        )
    finally:
        wb.close()


def analyze_max_mises_strain(path: str = None, reference_index: int = 0,
                             progress_cb=None) -> DicFieldResult:
    """[兼容包装] 等价于 analyze_max_field(path, 'mises', ...)。

    Args:
        path / reference_index / progress_cb: 同 analyze_max_field。

    Returns:
        DicFieldResult (metric_key='mises')

    Side Effects:
        见 analyze_max_field。
    """
    return analyze_max_field(path, 'mises', reference_index, progress_cb)


def _as_float(v) -> Optional[float]:
    """把单元格值安全转为 float; None 或非数值返回 None。

    Args:
        v: 单元格原始值 (可能为 None / 数字 / 字符串)

    Returns:
        float 或 None

    Side Effects:
        无 (纯函数)
    """
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_cylindrical(x, y, z=0.0):
    """把笛卡尔坐标 (x, y, z) 转换为圆柱坐标 (r, θ, z)。

    Args:
        x: X 坐标 (mm); None 视为无效
        y: Y 坐标 (mm); None 视为无效
        z: Z 坐标 (mm), 即圆柱坐标的轴向分量; None 按 0 处理

    Returns:
        (r, theta_deg, z) 三元组:
            - r:         径向距离 = sqrt(x^2 + y^2), 单位 mm
            - theta_deg: 极角 = atan2(y, x), 单位 度 (范围 -180~180)
            - z:         轴向坐标 (mm), 原样返回
        若 x 或 y 为 None, 返回 (None, None, z)。

    Side Effects:
        无 (纯函数)

    算法不变量:
        - 圆柱轴心取坐标系原点 (0, 0); 与本 DIC 数据的"径向距离/径向角"列定义
          一致 (该数据中 r=sqrt(x^2+y^2)、θ=atan2(y,x) 与导出列逐点吻合)。
        - θ 用度而非弧度, 与数据中 '径向角[deg]' 的单位一致。
    """
    if z is None:
        z = 0.0
    if x is None or y is None:
        return (None, None, z)
    r = math.hypot(float(x), float(y))
    theta_deg = math.degrees(math.atan2(float(y), float(x)))
    return (r, theta_deg, z)


def _stage_index(sheet_name: str) -> int:
    """从工作表名 (如 'STAGE427') 解析出末尾的时刻序号 427。

    Args:
        sheet_name: 工作表名

    Returns:
        解析得到的整数序号; 无法解析时返回 -1。

    Side Effects:
        无 (纯函数)

    算法不变量:
        - 仅截取名称末尾连续数字段; 'STAGE000' -> 0, 'STAGE500' -> 500。
    """
    digits = ''
    for ch in reversed(sheet_name):
        if ch.isdigit():
            digits = ch + digits
        else:
            break
    return int(digits) if digits else -1


# =============================================================================
# 报告格式化 (命令行用; GUI 改为内联展示, 不调用本函数)
# =============================================================================
def format_report(result: DicFieldResult) -> str:
    """把分析结果渲染为精简的多行中文文本报告 (命令行打印用)。

    仅保留: 数据文件、最大值、出现时刻、测点网格、坐标。

    Args:
        result: analyze_max_field 的返回值

    Returns:
        多行字符串报告

    Side Effects:
        无 (纯字符串拼装)
    """
    gm = result.global_max
    u = result.unit
    lines = []
    lines.append(f'===== DIC 全场最大 {result.metric_label} =====')
    lines.append(f'数据文件: {result.file_path}')
    if gm is None:
        lines.append('未找到任何有效数据 (全部为空)。')
        return '\n'.join(lines)

    lines.append(f'最大 {result.metric_label} : {gm.value:.6g} {u}')
    lines.append(f'出现时刻      : {gm.stage_name} (第 {gm.stage_index} 时刻)')
    lines.append(f'测点网格      : 行ROW={gm.grid_row}, 列COL={gm.grid_col}')
    if gm.ref_x_mm is not None:
        lines.append(f'参考态坐标(mm): X={_fmt(gm.ref_x_mm)}, Y={_fmt(gm.ref_y_mm)}')
        r, th, z = to_cylindrical(gm.ref_x_mm, gm.ref_y_mm, gm.ref_z_mm)
        lines.append(f'  圆柱坐标(参考): r={_fmt(r)} mm, θ={_fmt(th)}°, z={_fmt(z)} mm')
    lines.append(f'变形后坐标(mm): X={_fmt(gm.x_mm)}, Y={_fmt(gm.y_mm)}')
    r, th, z = to_cylindrical(gm.x_mm, gm.y_mm, gm.z_mm)
    lines.append(f'  圆柱坐标(变形后): r={_fmt(r)} mm, θ={_fmt(th)}°, z={_fmt(z)} mm')

    # 截面平均 (跨全部时刻取最大)
    if result.circ_theta_deg is not None:
        cs = result.circ_section
        if cs is not None:
            lines.append(f'最大平均周向Mises应变: {cs.value:.6g} {u} @ {cs.stage_name} '
                         f'({cs.param_desc}, n={cs.n_points})')
        else:
            lines.append(f'最大平均周向Mises应变: 子午截面 θ={result.circ_theta_deg:.1f}° 无数据点')
    if result.radial_radius is not None:
        rs = result.radial_section
        if rs is not None:
            lines.append(f'最大平均径向Mises应变: {rs.value:.6g} {u} @ {rs.stage_name} '
                         f'({rs.param_desc}, n={rs.n_points})')
        else:
            lines.append(f'最大平均径向Mises应变: 圆柱截面 r={result.radial_radius:.0f} 无数据点')
    if result.skipped_stages:
        lines.append(f'已跳过 {len(result.skipped_stages)} 个含异常数据的时刻表: '
                     f'{", ".join(result.skipped_stages)}')
    return '\n'.join(lines)


def _fmt(v: Optional[float]) -> str:
    """坐标/数值格式化为 3 位小数字符串; None 显示 'N/A'。

    Args:
        v: 数值或 None

    Returns:
        格式化字符串

    Side Effects:
        无 (纯函数)
    """
    return 'N/A' if v is None else f'{v:.3f}'


# =============================================================================
# 命令行入口
# =============================================================================
def main(argv=None) -> int:
    """命令行入口: 分析 DIC 文件并打印全场最大 Mises 等效应变报告。

    Args:
        argv: 参数列表 (默认 sys.argv[1:]):
              [0] xlsx 路径 (省略=自动在 DICdata 查找)
              [1] 圆柱截面半径 R (mm, 省略=不算平均径向)
              [2] 子午截面角度 θ (度, 省略=不算平均周向)

    Returns:
        进程退出码: 0 成功; 非 0 表示出错。

    Side Effects:
        - 向标准输出打印进度与报告。
    """
    args = sys.argv[1:] if argv is None else argv
    path = args[0] if len(args) >= 1 else None
    radius = float(args[1]) if len(args) >= 2 else None
    theta = float(args[2]) if len(args) >= 3 else None

    def _cli_progress(done: int, total: int) -> bool:
        if done == total or done % 25 == 0:
            print(f'\r  进度: {done}/{total} 时刻', end='', flush=True)
        return True

    try:
        result = analyze_max_field(path, 'mises', progress_cb=_cli_progress,
                                   circ_theta_deg=theta, radial_radius=radius)
    except Exception as ex:  # noqa: BLE001 - CLI 友好提示
        print(f'\n[错误] {ex}')
        return 1
    print()  # 进度行换行
    print(format_report(result))
    return 0


if __name__ == '__main__':
    sys.exit(main())
